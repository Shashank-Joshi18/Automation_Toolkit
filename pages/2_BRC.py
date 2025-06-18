import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import os

def analyze_roles(file_bytes: BytesIO) -> BytesIO:
    # Read Excel file and normalize column names
    df = pd.read_excel(file_bytes)
    df.columns = df.columns.str.strip().str.lower()

    required_columns = ["business profile", "name", "it application", "access right"]
    for col in required_columns:
        if col not in df.columns:
            raise ValueError(f"Missing required column: '{col}'")

    # Clean up data and normalize role info
    for col in required_columns:
        df[col] = df[col].astype(str).str.strip()

    df['role'] = (
        df['it application'].astype(str).str.strip().str.lower() +
        " - " +
        df['access right'].astype(str).str.strip().str.lower()
    )

    # Safety check for missing values
    if df[required_columns].isnull().any().any():
        raise ValueError("The file contains missing values in required columns. Please check the data.")

    job_roles = df['business profile'].unique()
    job_role_user_roles = {}
    role_to_users_per_job = {}

    for job in job_roles:
        job_df = df[df['business profile'] == job]
        user_roles = {}

        for _, row in job_df.iterrows():
            user = row['name']
            role = row['role']
            user_roles.setdefault(user, set()).add(role)
            role_to_users_per_job.setdefault(job, {}).setdefault(role, set()).add(user)

        all_roles = set.union(*user_roles.values()) if user_roles else set()
        common_roles = set.intersection(*user_roles.values()) if user_roles else set()
        uncommon_roles = all_roles - common_roles

        job_role_user_roles[job] = {
            "common": common_roles,
            "uncommon": uncommon_roles,
            "role_to_users": role_to_users_per_job[job]
        }

    job_common_sets = [v['common'] for v in job_role_user_roles.values() if v['common']]
    universal_common_roles = set.intersection(*job_common_sets) if job_common_sets else set()

    def format_roles(role_set, role_to_users, include_users=True):
        data = []
        for role in sorted(role_set):
            if pd.isna(role):
                continue

            if isinstance(role, str) and " - " in role:
                app, role_name = role.split(" - ", 1)
            else:
                app, role_name = "Unknown", str(role)

            row = [app.strip(), role_name.strip()]
            assigned_to = ""

            if include_users:
                users = role_to_users.get(role, [])
                assigned_to = ", ".join(sorted(str(u) for u in users if pd.notna(u)))

            row += [assigned_to, "", ""]  # Description, Keep/Revoke
            data.append(row)

        columns = ["IT Application", "Role", "Assigned To", "Description", "Keep/Revoke"]
        return pd.DataFrame(data, columns=columns)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        original_df = df.drop(columns=['role'])
        original_df.to_excel(writer, sheet_name="Original Data", index=False)

        universal_df = format_roles(universal_common_roles, {}, include_users=False)
        universal_df.to_excel(writer, sheet_name="Universal Common Roles", index=False)

        for job, role_data in job_role_user_roles.items():
            clean_job = job.replace("/", "_")
            job_specific_common = role_data['common'] - universal_common_roles
            common_df = format_roles(job_specific_common, role_data['role_to_users'], include_users=False)
            uncommon_df = format_roles(role_data['uncommon'], role_data['role_to_users'], include_users=True)

            common_df.to_excel(writer, sheet_name=f"{clean_job} - Common", index=False)
            uncommon_df.to_excel(writer, sheet_name=f"{clean_job} - Uncommon", index=False)

    # Apply formatting
    output.seek(0)
    wb = load_workbook(output)

    def style_headers(ws, fill_color):
        for cell in ws[1]:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in ws.columns:
            try:
                values = [len(str(cell.value)) for cell in col[1:] if cell.value is not None]
                max_len = max(values) if values else 12
            except Exception:
                max_len = 12
            adjusted_width = min(max(max_len + 2, 12), 40)
            ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

        ws.row_dimensions[1].height = 25

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        color = "4472C4" if "Original" in sheet_name else "228B22"
        style_headers(ws, color)

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output

# ---------- STREAMLIT UI ----------

st.set_page_config(page_title="BRC", layout="centered")
st.title("🔍 Common vs Uncommon Role Analyzer")

uploaded_file = st.file_uploader("📤 Upload the Excel File", type=["xlsx"])

if uploaded_file:
    st.success("✅ File uploaded.")

    if st.button("Analyze Roles"):
        try:
            result_file = analyze_roles(uploaded_file)
            input_name = os.path.splitext(uploaded_file.name)[0]
            revised_filename = f"{input_name}_revised.xlsx"

            st.success("✅ Analysis complete. Download the result:")
            st.download_button(
                label="📥 Download Updated Excel File",
                data=result_file,
                file_name=revised_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"❌ Error: {e}")
